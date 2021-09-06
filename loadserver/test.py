# -- coding:utf-8 --
import os
import socket
import uuid
import base64
from osgeo import ogr
from osgeo import osr
from osgeo import gdal
import sys
import csv
import json
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from openpyxl import load_workbook
import pymysql
import math
# iobjectspy.set_iobjects_java_path('E:\SuperMapDemo\yaogan\javaforobject\Bin')


maxInt = sys.maxsize
decrement = True
while decrement:
    decrement = False
    try:
        csv.field_size_limit(maxInt)
    except OverflowError:
        maxInt = int(maxInt/10)
        decrement = True
# BP能源导入
def test_import_excel():
    conn = pymysql.connect(
        host="192.168.84.130",
        port=3306,
        user="root",
        password="0",
        database="dilisuo",
        charset="utf8"
    )
    # 获取操作游标
    cursor = conn.cursor()
    dir_path = "D:/SuperMapDemo/地理所/【中科院】第二次数据对接20210430/【0】BP能源属性数据_任亚文/能源属性数据/"
    # 获取所有文件名
    file_names = os.listdir(dir_path)
    for x in range(len(file_names)):
        # if file_names[x] == "可再生-风能（装机容量）.xlsx":
            file_path = dir_path+file_names[x]
            file = load_workbook(file_path)
            table = file.get_sheet_by_name(file.sheetnames[0])
            # 行数
            rows = table.max_row
            # 列数
            cols = table.max_column
            temp_values=[]
            temp_year=[]
            temp_unit = ""
            temp_type = ""
            for row in range(1,rows+1):
                temp_value = {}
                temp_value["year"] = []
                temp_value["values"] = []
                for col in range(1, cols + 1):
                    value = table.cell(row=row, column=col).value
                    if row == 1 and col == 1:
                        temp_unit = value
                        temp_type = file_names[x].split(".")[0]
                    elif row == 1 and col != 1:
                        if value != "" and value != None:
                            temp_year.append(value)
                    elif row !=1 and col == 1:
                        if value != "" and value != None:
                            temp_value["country_en"] = value
                    elif row !=1 and col != 1:
                        if value == "n/a":
                            temp_value["values"].append(0)
                            temp_value["year"].append(temp_year[col - 2])
                        elif value !="" and value!=None:
                            temp_value["values"].append(value)
                            temp_value["year"].append(temp_year[col - 2])
                if len(temp_value["values"]) != 0 and "country_en" in temp_value.keys():
                    temp_value["unit"] = temp_unit
                    temp_value["type"] = temp_type
                    temp_values.append(temp_value)
            for i in range(len(temp_values)):
                for j in range(len(temp_values[i]["year"])):
                    sql_str = "insert into bpnengyuan (country_en,type,year,unit,value) values \
                                     ('%s','%s','%s','%s','%s')" % \
                              (temp_values[i]["country_en"],temp_values[i]["type"],temp_values[i]["year"][j],\
                               temp_values[i]["unit"],temp_values[i]["values"][j])
                    cursor.execute(sql_str)
                    conn.commit()
    conn.close()
# 能源贸易
def test_import_excel2():
    conn = pymysql.connect(
        host="192.168.84.130",
        port=3306,
        user="root",
        password="0",
        database="dilisuo",
        charset="utf8"
    )
    # 获取操作游标
    cursor = conn.cursor()
    dir_path = "D:/SuperMapDemo/地理所/【中科院】第二次数据对接20210430/【1】能源贸易_何则/"
    # 获取所有文件名
    file_names = os.listdir(dir_path)
    for x in range(len(file_names)):
        if file_names[x].find("xlsx") !=-1:
            print(file_names[x])
            file_path = dir_path+file_names[x]
            file = load_workbook(file_path)
            table = file.get_sheet_by_name(file.sheetnames[0])
            # 行数
            rows = table.max_row
            # 列数
            cols = table.max_column
            temp_values=[]
            for row in range(1,rows+1):
                temp_value = {}
                for col in range(1, cols + 1):
                    value = table.cell(row=row, column=col).value
                    if row != 1 and col == 1:
                        if value == None or value == "":
                            temp_value["exporter_iso3"] = ""
                        else:
                            temp_value["exporter_iso3"] = value
                    elif row != 1 and col == 2:
                        if value == None or value == "":
                            temp_value["exporter"] = ""
                        else:
                            temp_value["exporter"] = value
                    elif row !=1 and col == 3:
                        if value == None or value == "":
                            temp_value["importer_iso3"] = ""
                        else:
                            temp_value["importer_iso3"] = value
                    elif row !=1 and col == 4:
                        if value == None or value == "":
                            temp_value["importer"] = ""
                        else:
                            temp_value["importer"] = value

                    elif row !=1 and col == 5:
                        if value == None or value == "":
                            temp_value["type"] = ""
                        else:
                            temp_value["type"] = value
                    elif row !=1 and col == 6:
                        if value == None or value == "":
                            temp_value["code"] = 0
                        else:
                            temp_value["code"] = value
                    elif row !=1 and col == 7:
                        if value == None or value == "":
                            temp_value["year"] = ""
                        else:
                            temp_value["year"] = value
                    elif row !=1 and col == 8:
                        if value == "" or value == None:
                            temp_value["value"] = 0
                        else:
                            temp_value["value"] = value
                    elif row !=1 and col == 9:
                        if value == "" or value == None:
                            temp_value["weight"] = 0
                        else:
                            temp_value["weight"] = value
                if "exporter_iso3" in temp_value.keys():
                    temp_value["value_unit"] = "1000USD"
                    temp_value["weight_unit"] = "1000kg"
                    temp_values.append(temp_value)
            for i in range(len(temp_values)):
                sql_str = "insert into nengyuanmaoyi (exporter_iso3,exporter,importer_iso3,importer,type,code,year,value,value_unit,weight,weight_unit) values \
                                 ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" % \
                          (temp_values[i]["exporter_iso3"],temp_values[i]["exporter"],temp_values[i]["importer_iso3"],\
                           temp_values[i]["importer"],temp_values[i]["type"],temp_values[i]["code"],temp_values[i]["year"],\
                           temp_values[i]["value"],temp_values[i]["value_unit"],temp_values[i]["weight"],temp_values[i]["weight_unit"])
                cursor.execute(sql_str)
                conn.commit()
    conn.close()
# 新能源贸易导入
def test_import_excel3():
    conn = pymysql.connect(
        host="192.168.84.130",
        port=3306,
        user="root",
        password="0",
        database="dilisuo",
        charset="utf8"
    )
    # 获取操作游标
    cursor = conn.cursor()
    dir_path = "D:/SuperMapDemo/地理所/【中科院】第二次数据对接20210430/【2】新能源贸易_夏四友/15种新能源关键稀缺材料/"
    # 获取所有文件名
    file_names = os.listdir(dir_path)
    for x in range(len(file_names)):
        # 拼接地址
        temp_path = dir_path+file_names[x]
        # 判断是否是目录
        if os.path.isdir(temp_path):
            temp_file_names = os.listdir(temp_path)
            for s in range(len(temp_file_names)):
                temp_file_path = temp_path+"/"+temp_file_names[s]
                file = load_workbook(temp_file_path)
                table = file.get_sheet_by_name(file.sheetnames[0])
                # 行数
                rows = table.max_row
                # 列数
                cols = table.max_column
                temp_values=[]
                for row in range(1,rows+1):
                    temp_value = {}
                    for col in range(1, cols + 1):
                        value = table.cell(row=row, column=col).value
                        if row != 1 and col == 1:
                            if value == None or value == "":
                                temp_value["exporter_iso3"] = ""
                            else:
                                temp_value["exporter_iso3"] = value
                        elif row != 1 and col == 2:
                            if value == None or value == "":
                                temp_value["exporter"] = ""
                            elif value.find("Ivoire")!=-1:
                                temp_value["exporter"] = value.replace("'"," ")
                            else:
                                temp_value["exporter"] = value
                        elif row !=1 and col == 3:
                            if value == None or value == "":
                                temp_value["importer_iso3"] = ""
                            else:
                                temp_value["importer_iso3"] = value
                        elif row !=1 and col == 4:
                            if value == None or value == "":
                                temp_value["importer"] = ""
                            elif value.find("Ivoire")!=-1:
                                temp_value["importer"] = value.replace("'"," ")
                            else:
                                temp_value["importer"] = value

                        elif row !=1 and col == 5:
                            if value == None or value == "":
                                temp_value["type"] = ""
                            else:
                                temp_value["type"] = value
                        elif row !=1 and col == 6:
                            if value == None or value == "":
                                temp_value["code"] = 0
                            else:
                                temp_value["code"] = value
                        elif row !=1 and col == 7:
                            if value == None or value == "":
                                temp_value["year"] = ""
                            else:
                                temp_value["year"] = value
                        elif row !=1 and col == 8:
                            if value == "" or value == None:
                                temp_value["value"] = 0
                            else:
                                temp_value["value"] = value
                        elif row !=1 and col == 9:
                            if value == "" or value == None:
                                temp_value["weight"] = 0
                            else:
                                temp_value["weight"] = value
                    if "exporter" in temp_value.keys() and "importer" in temp_value.keys():
                        if temp_value["exporter"] != "" and temp_value["importer"] != "":
                            temp_value["value_unit"] = "1000USD"
                            temp_value["weight_unit"] = "1000kg"
                            temp_values.append(temp_value)
                for i in range(len(temp_values)):
                    sql_str = "insert into xinnengyuanmaoyi (exporter_iso3,exporter,importer_iso3,importer,type,code,year,value,value_unit,weight,weight_unit) values \
                                     ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" % \
                              (temp_values[i]["exporter_iso3"],temp_values[i]["exporter"],temp_values[i]["importer_iso3"],\
                               temp_values[i]["importer"],temp_values[i]["type"],temp_values[i]["code"],temp_values[i]["year"],\
                               temp_values[i]["value"],temp_values[i]["value_unit"],temp_values[i]["weight"],temp_values[i]["weight_unit"])
                    cursor.execute(sql_str)
                    conn.commit()
                print(temp_file_path+"插入成功")
    conn.close()
# 隐含能源贸易导入 目前只有煤炭
def test_import_excel4():
    conn = pymysql.connect(
        host="192.168.84.130",
        port=3306,
        user="root",
        password="0",
        database="dilisuo",
        charset="utf8"
    )
    # 获取操作游标
    cursor = conn.cursor()
    dir_path = "D:/SuperMapDemo/地理所/【中科院】第二次数据对接20210430/【5】隐含能源贸易_周彦楠/【1】隐含能源OD数据/【1】coal/"
    # 获取所有文件名
    file_names = os.listdir(dir_path)
    for x in range(len(file_names)):
        temp_file_path = dir_path+file_names[x]
        file = load_workbook(temp_file_path)
        table = file.get_sheet_by_name(file.sheetnames[1])
        # 行数
        rows = table.max_row
        # 列数
        cols = table.max_column
        temp_values=[]
        for row in range(1,rows+1):
            temp_value = {}
            for col in range(1, cols + 1):
                value = table.cell(row=row, column=col).value
                if row != 1 and col == 1:
                    if value == None or value == "":
                        temp_value["exporter"] = ""
                    elif value.find("Ivoire") != -1:
                        temp_value["exporter"] = value.replace("'", " ")
                    else:
                        temp_value["exporter"] = value
                elif row != 1 and col == 2:
                    if value == None or value == "":
                        temp_value["importer"] = ""
                    elif value.find("Ivoire")!=-1:
                        temp_value["importer"] = value.replace("'"," ")
                    else:
                        temp_value["importer"] = value
                elif row !=1 and col == 3:
                    if value == None or value == "":
                        temp_value["weight"] = 0
                    else:
                        temp_value["weight"] = value
            if "exporter" in temp_value.keys() and "importer" in temp_value.keys():
                if temp_value["exporter"] != "" and temp_value["importer"] != "":
                    temp_value["weight_unit"] = "mtoe"
                    temp_value["type"] = "coal"
                    temp_value["year"] = file_names[x].split("_")[0]
                    temp_values.append(temp_value)
        print(temp_values)
        for i in range(len(temp_values)):
            sql_str = "insert into yinhannengyuanmaoyi (exporter,importer,type,year,weight,weight_unit) values \
                             ('%s','%s','%s','%s','%s','%s')" % \
                      (temp_values[i]["exporter"],temp_values[i]["importer"],temp_values[i]["type"],temp_values[i]["year"],\
                       temp_values[i]["weight"],temp_values[i]["weight_unit"])
            cursor.execute(sql_str)
            conn.commit()
        print(temp_file_path+"插入成功")
    conn.close()
# 隐含能源贸易分布 目前仅有煤炭
def test_import_excel5():
    conn = pymysql.connect(
        host="192.168.84.130",
        port=3306,
        user="root",
        password="0",
        database="dilisuo",
        charset="utf8"
    )
    # 获取操作游标
    cursor = conn.cursor()
    dir_path = "D:/SuperMapDemo/地理所/【中科院】第二次数据对接20210430/【5】隐含能源贸易_周彦楠/【2】隐含能源分布数据/"
    # 获取所有文件名
    file_names = os.listdir(dir_path)
    for x in range(len(file_names)):
        temp_file_path = dir_path+file_names[x]
        file = load_workbook(temp_file_path)
        table = file.get_sheet_by_name(file.sheetnames[0])
        # 行数
        rows = table.max_row
        # 列数
        cols = table.max_column
        temp_values=[]
        for row in range(1,rows+1):
            temp_value = {}
            for col in range(1, cols + 1):
                value = table.cell(row=row, column=col).value
                if row != 1 and col == 1:
                    if value == None or value == "":
                        temp_value["country"] = ""
                    elif value.find("Ivoire") != -1:
                        temp_value["country"] = value.replace("'", " ")
                    else:
                        temp_value["country"] = value
                if row != 1 and col == 2:
                    if value == None or value == "":
                        temp_value["export_weight"] = 0
                    else:
                        temp_value["export_weight"] = value
                elif row != 1 and col == 3:
                    if value == None or value == "":
                        temp_value["import_weight"] = 0
                    else:
                        temp_value["import_weight"] = value
                elif row !=1 and col == 4:
                    if value == None or value == "":
                        temp_value["net_export_weight"] = 0
                    else:
                        temp_value["net_export_weight"] = value
            if "country" in temp_value.keys():
                if temp_value["country"] != "":
                    temp_value["weight_unit"] = "mtoe"
                    temp_value["type"] = "coal"
                    temp_value["year"] = file_names[x].split(".")[0].split("出口")[1]
                    temp_values.append(temp_value)
        for i in range(len(temp_values)):
            sql_str = "insert into yinhannengyuanmaoyifenbu (country,export_weight,import_weight,net_export_weight,weight_unit,type,year) values \
                             ('%s','%s','%s','%s','%s','%s','%s')" % \
                      (temp_values[i]["country"],temp_values[i]["export_weight"],temp_values[i]["import_weight"],temp_values[i]["net_export_weight"],\
                       temp_values[i]["weight_unit"],temp_values[i]["type"],temp_values[i]["year"])
            cursor.execute(sql_str)
            conn.commit()
        print(temp_file_path+"插入成功")
        # 批量插入
        # sql_str = "insert into yinhannengyuanmaoyifenbu (country,export_weight,import_weight,net_export_weight,weight_unit,type,year) values \
        #                  (%s,%s,%s,%s,%s,%s,%s)"
        # res = cursor.executemany(sql_str,temp_values)
        # print(res)
        # conn.commit()
        # print(temp_file_path+"插入成功")
    conn.close()
#专利转移数据导入
def test_import_excel6():
    conn = pymysql.connect(
        host="192.168.84.130",
        port=3306,
        user="root",
        password="0",
        database="dilisuo",
        charset="utf8"
    )
    # 获取操作游标
    cursor = conn.cursor()
    dir_path = "D:/SuperMapDemo/地理所/【中科院】第二次数据对接20210430/【6】能源技术_钱肖颖/"
    # 获取所有文件名
    file_names = os.listdir(dir_path)
    for x in range(len(file_names)):
        if file_names[x].find("xlsx") != -1 and file_names[x].find("~$") == -1:
            temp_file_path = dir_path+file_names[x]
            file = load_workbook(temp_file_path)
            table = file.get_sheet_by_name(file.sheetnames[0])
            # 行数
            rows = table.max_row
            # 列数
            cols = table.max_column
            temp_values=[]
            for row in range(1,rows+1):
                temp_value = {}
                for col in range(1, cols + 1):
                    value = table.cell(row=row, column=col).value
                    if row != 1 and row !=2 and col == 1:
                        temp_value["year"] = str(value)
                    elif row != 1 and row !=2 and col == 2:
                        temp_value["title_en"] = str(value)
                    elif row != 1 and row !=2 and col == 3:
                        temp_value["abstract_en"] = str(value)
                    elif row != 1 and row !=2 and col == 4:
                        temp_value["pub_year"] = str(value)
                    elif row != 1 and row !=2 and col == 5:
                        temp_value["title_ch"] = str(value)
                    elif row != 1 and row !=2 and col == 6:
                        temp_value["abstract_ch"] = str(value)
                    elif row != 1 and row !=2 and col == 7:
                        temp_value["main_ipc"] = str(value)
                    elif row != 1 and row !=2 and col == 8:
                        temp_value["ipc"] = str(value)
                    elif row != 1 and row !=2 and col == 9:
                        if value == None or value == "":
                            temp_value["exporter"] = ""
                        else:
                            temp_value["exporter"] = str(value)
                    elif row != 1 and row !=2 and col == 10:
                        if value == None or value == "":
                            temp_value["importer"] = ""
                        else:
                            temp_value["importer"] = str(value)
                if "year" in temp_value.keys():
                    temp_values.append(temp_value)
            for i in range(len(temp_values)):
                sql_str = "insert into zhuanlizhuanyi (year,title_en,abstract_en,pub_year,title_ch,abstract_ch,main_ipc,ipc,exporter,importer) values \
                                 ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" % \
                          (temp_values[i]["year"], temp_values[i]["title_en"], temp_values[i]["abstract_en"],\
                           temp_values[i]["pub_year"], temp_values[i]["title_ch"],\
                           temp_values[i]["abstract_ch"], temp_values[i]["main_ipc"], temp_values[i]["ipc"],\
                           temp_values[i]["exporter"],temp_values[i]["importer"])
                cursor.execute(sql_str)
                conn.commit()
            print(temp_file_path + "插入成功")
    conn.close()
#能源治理组织导入
def test_import_excel7():
    conn = pymysql.connect(
        host="192.168.84.130",
        port=3306,
        user="root",
        password="0",
        database="dilisuo",
        charset="utf8"
    )
    # 获取操作游标
    cursor = conn.cursor()
    dir_path = "D:/SuperMapDemo/地理所/【中科院】第二次数据对接20210430/【7】能源组织_任亚文/"
    # 获取所有文件名
    file_names = os.listdir(dir_path)
    for x in range(len(file_names)):
        if file_names[x].find("xlsx") != -1 and file_names[x].find("~$") == -1:
            temp_file_path = dir_path+file_names[x]
            file = load_workbook(temp_file_path)
            table = file.get_sheet_by_name(file.sheetnames[0])
            # 行数
            rows = table.max_row
            # 列数
            cols = table.max_column
            temp_values=[]
            # 组织列表
            temp_organizations=[]
            for row in range(1,rows+1):
                temp_value = {}
                temp_value["is_joins"]=[]
                for col in range(1, cols + 1):
                    value = table.cell(row=row, column=col).value
                    if row == 1 and col > 2:
                        if value !="" and value !=None:
                            temp_organizations.append(str(value))
                    elif row != 1 and col == 1:
                        temp_value["country_ch"] = str(value)
                    elif row != 1 and col == 2:
                        temp_value["country_en"] = str(value)
                    elif row != 1 and col > 2:
                        if value != "" and value != None and str(value).find("=SUM")==-1:
                            temp_value["is_joins"].append(str(value))
                if len(temp_value["is_joins"]) !=0:
                    temp_value["organizations"] = temp_organizations
                    temp_values.append(temp_value)
            for i in range(len(temp_values)):
                for j in range(len(temp_values[i]["organizations"])):
                    sql_str = "insert into nengyuanzhilizuzhi (country_ch,country_en,organization,is_join) values \
                                     ('%s','%s','%s','%s')" % \
                              (temp_values[i]["country_ch"], temp_values[i]["country_en"], temp_values[i]["organizations"][j],\
                               temp_values[i]["is_joins"][j])
                    cursor.execute(sql_str)
                    conn.commit()
            print(temp_file_path + "插入成功")
    conn.close()

#能源企业导入 目前只有中国
def test_import_excel8():
    conn = pymysql.connect(
        host="192.168.84.130",
        port=3306,
        user="root",
        password="0",
        database="dilisuo",
        charset="utf8"
    )
    # 获取操作游标
    cursor = conn.cursor()
    dir_path = "D:/SuperMapDemo/地理所/【中科院】第二次数据对接20210430/【11】中国能源企业的点位数据_钱肖颖/"
    # 获取所有文件名
    file_names = os.listdir(dir_path)
    for x in range(len(file_names)):
        if file_names[x].find("xlsx") != -1 and file_names[x].find("~$") == -1:
            temp_file_path = dir_path+file_names[x]
            file = load_workbook(temp_file_path)
            table = file.get_sheet_by_name(file.sheetnames[0])
            # 行数
            rows = table.max_row
            # 列数
            cols = table.max_column
            temp_values=[]
            for row in range(1,rows+1):
                temp_value = {}
                for col in range(1, cols + 1):
                    value = table.cell(row=row, column=col).value
                    if row != 1 and row !=2 and col == 1:
                        temp_value["company"] = str(value)
                    elif row != 1 and row !=2 and col == 2:
                        temp_value["registration"] = str(value)
                    elif row != 1 and row !=2 and col == 3:
                        temp_value["capital"] = str(value)
                    elif row != 1 and row !=2 and col == 4:
                        if value == None:
                            temp_value["Incorporation_year"] = ""
                        else:
                            temp_value["Incorporation_year"] = str(value)
                    elif row != 1 and row !=2 and col == 5:
                        if value == None:
                            temp_value["approval_year"] = ""
                        else:
                            temp_value["approval_year"] = str(value)
                    elif row != 1 and row !=2 and col == 6:
                        temp_value["province"] = str(value)
                    elif row != 1 and row !=2 and col == 7:
                        temp_value["city"] = str(value)
                    elif row != 1 and row !=2 and col == 8:
                        temp_value["county"] = str(value)
                    elif row != 1 and row !=2 and col == 9:
                        temp_value["type"] = str(value)
                    elif row != 1 and row !=2 and col == 10:
                        temp_value["sector"] = str(value)
                    elif row != 1 and row !=2 and col == 11:
                        temp_value["address"] = str(value)
                    elif row != 1 and row !=2 and col == 12:
                        temp_value["business_scope"] = str(value)
                    elif row != 1 and row !=2 and col == 13:
                        temp_value["lon"] = str(value)
                    elif row != 1 and row !=2 and col == 14:
                        temp_value["lat"] = str(value)
                if "company" in temp_value.keys():
                    temp_values.append(temp_value)
            print(temp_values)
            for i in range(len(temp_values)):
                sql_str = "insert into nengyuanqiye (company,registration,capital,Incorporation_year,approval_year,province,city,county,\
                            type,sector,address,business_scope,lon,lat) values \
                                 ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" % \
                          (temp_values[i]["company"], temp_values[i]["registration"], temp_values[i]["capital"],\
                           temp_values[i]["Incorporation_year"], temp_values[i]["approval_year"],\
                           temp_values[i]["province"], temp_values[i]["city"], temp_values[i]["county"],\
                           temp_values[i]["type"],temp_values[i]["sector"],temp_values[i]["address"],temp_values[i]["business_scope"],\
                           temp_values[i]["lon"],temp_values[i]["lat"])
                cursor.execute(sql_str)
                conn.commit()
            print(temp_file_path + "插入成功")
    conn.close()
#全球电力装机容量 导入
def test_import_excel9():
    conn = pymysql.connect(
        host="192.168.84.130",
        port=3306,
        user="root",
        password="0",
        database="dilisuo",
        charset="utf8"
    )
    # 获取操作游标
    cursor = conn.cursor()
    dir_path = "D:/SuperMapDemo/地理所/【中科院】第二次数据对接20210430/【10】全球能源基础设施数据_何则&周彦楠/【2】全球电力装机容量/"
    # 获取所有文件名
    file_names = os.listdir(dir_path)
    for x in range(len(file_names)):
        if file_names[x].find("xlsx") != -1 and file_names[x].find("~$") == -1:
            temp_file_path = dir_path+file_names[x]
            file = load_workbook(temp_file_path)
            table = file.get_sheet_by_name(file.sheetnames[0])
            # 行数
            rows = table.max_row
            # 列数
            cols = table.max_column
            temp_values=[]
            for row in range(1,rows+1):
                temp_value = {}
                for col in range(1, cols + 1):
                    value = table.cell(row=row, column=col).value
                    if row != 1 and col == 1:
                        temp_value["year"] = str(value)
                    elif row != 1 and col == 2:
                        temp_value["weight"] = str(value)
                    elif row != 1 and col == 3:
                        temp_value["country"] = str(value)
                    elif row != 1 and col == 4:
                        temp_value["country_iso3"] = str(value)
                    elif row != 1 and col == 5:
                        temp_value["weight_unit"] = str(value)
                if "year" in temp_value.keys():
                    temp_values.append(temp_value)
            print(temp_values)
            for i in range(len(temp_values)):
                sql_str = "insert into dianlizhuangjirongliang (year,weight,country,country_iso3,weight_unit) values \
                                 ('%s','%s','%s','%s','%s')" % \
                          (temp_values[i]["year"], temp_values[i]["weight"], temp_values[i]["country"],\
                           temp_values[i]["country_iso3"], temp_values[i]["weight_unit"])
                cursor.execute(sql_str)
                conn.commit()
            print(temp_file_path + "插入成功")
    conn.close()
# 导入船舶轨迹数据
def test_import_excel10():
    conn = pymysql.connect(
        host="192.168.84.130",
        port=3306,
        user="root",
        password="0",
        database="dilisuo",
        charset="utf8"
    )
    # 获取操作游标
    cursor = conn.cursor()
    dir_path = "D:/SuperMapDemo/地理所/【中科院】第二次数据对接20210430/【8】AIS的传播轨迹数据_彭澎/"
    # 获取所有文件名
    file_names = os.listdir(dir_path)
    for x in range(len(file_names)):
        if file_names[x].find("xlsx") != -1 and file_names[x].find("~$") == -1:
            temp_file_path = dir_path+file_names[x]
            file = load_workbook(temp_file_path)
            table = file.get_sheet_by_name(file.sheetnames[0])
            # 行数
            rows = table.max_row
            # 列数
            cols = table.max_column
            temp_values=[]
            for row in range(1,rows+1):
                temp_value = {}
                for col in range(1, cols + 1):
                    value = table.cell(row=row, column=col).value
                    if row != 1 and row != 2 and col == 1:
                        temp_value["mmsi"] = str(value)
                    elif row != 1 and row != 2 and col == 2:
                        temp_value["lon"] = str(value)
                    elif row != 1 and row != 2 and col == 3:
                        temp_value["lat"] = str(value)
                    elif row != 1 and row != 2 and col == 4:
                        temp_value["time"] = str(value)
                if "mmsi" in temp_value.keys():
                    temp_values.append(temp_value)
            print(temp_values)
            for i in range(len(temp_values)):
                sql_str = "insert into chuanboguiji (mmsi,lon,lat,time) values \
                                 ('%s','%s','%s','%s')" % \
                          (temp_values[i]["mmsi"], temp_values[i]["lon"], temp_values[i]["lat"],\
                           temp_values[i]["time"])
                cursor.execute(sql_str)
                conn.commit()
            print(temp_file_path + "插入成功")
    conn.close()
# 导入油气并购数据
def test_import_excel11():
    conn = pymysql.connect(
        host="192.168.84.130",
        port=3306,
        user="root",
        password="0",
        database="dilisuo",
        charset="utf8"
    )
    # 获取操作游标
    cursor = conn.cursor()
    dir_path = "D:/SuperMapDemo/地理所/【中科院】第二次数据对接20210430/【4】能源投资并购_郭越/"
    # 获取所有文件名
    file_names = os.listdir(dir_path)
    for x in range(len(file_names)):
        if file_names[x].find("xlsx") != -1 and file_names[x].find("~$") == -1:
            temp_file_path = dir_path+file_names[x]
            file = load_workbook(temp_file_path)
            table = file.get_sheet_by_name(file.sheetnames[2])
            # 行数
            rows = table.max_row
            # 列数
            cols = table.max_column
            temp_values=[]
            for row in range(1,rows+1):
                temp_value = {}
                for col in range(1, cols + 1):
                    value = table.cell(row=row, column=col).value
                    if row != 1 and col == 1:
                        temp_value["time"] = str(value)
                    elif row != 1 and col == 2:
                        temp_value["acquiror_country"] = str(value)
                    elif row != 1 and col == 3:
                        temp_value["target_country"] = str(value)
                    elif row != 1 and col == 4:
                        temp_value["value"] = str(value)
                if "acquiror_country" in temp_value.keys() and temp_value["acquiror_country"] != None:
                    temp_value["value_unit"] = "EUR"
                    temp_values.append(temp_value)
            print(temp_values)
            for i in range(len(temp_values)):
                sql_str = "insert into youqibinggou (time,acquiror_country,target_country,value,value_unit) values \
                                 ('%s','%s','%s','%s','%s')" % \
                          (temp_values[i]["time"], temp_values[i]["acquiror_country"], temp_values[i]["target_country"],\
                           temp_values[i]["value"],temp_values[i]["value_unit"])
                cursor.execute(sql_str)
                conn.commit()
            print(temp_file_path + "插入成功")
    conn.close()
# 导入全球发电厂数据
def test_import_excel12():
    conn = pymysql.connect(
        host="192.168.84.130",
        port=3306,
        user="root",
        password="0",
        database="dilisuo",
        charset="utf8"
    )
    # 获取操作游标
    cursor = conn.cursor()
    dir_path = "D:/SuperMapDemo/地理所/【中科院】第二次数据对接20210430/【10】全球能源基础设施数据_何则&周彦楠/【1】全球发电厂数据/"
    # 获取所有文件名
    file_names = os.listdir(dir_path)
    for x in range(len(file_names)):
        if file_names[x].find("xlsx") != -1 and file_names[x].find("~$") == -1:
            temp_file_path = dir_path+file_names[x]
            file = load_workbook(temp_file_path)
            table = file.get_sheet_by_name(file.sheetnames[0])
            # 行数
            rows = table.max_row
            # 列数
            cols = table.max_column
            temp_values=[]
            for row in range(1,rows+1):
                temp_value = {}
                for col in range(1, cols + 1):
                    value = table.cell(row=row, column=col).value
                    if value == None:
                        value = ""
                    if row != 1 and row !=2 and col == 1:
                        temp_value["country_iso3"] = str(value)
                    elif row != 1 and row !=2 and col == 2:
                        temp_value["country"] = str(value)
                    elif row != 1 and row !=2 and col == 3:
                        temp_value["power_station_name"] = str(value)
                    elif row != 1 and row !=2 and col == 4:
                        temp_value["power_station_code"] = str(value)
                    elif row != 1 and row !=2 and col == 5:
                        temp_value["power_station_capacity"] = str(value)
                    elif row != 1 and row !=2 and col == 6:
                        temp_value["lat"] = str(value)
                    elif row != 1 and row !=2 and col == 7:
                        temp_value["lon"] = str(value)
                    elif row != 1 and row !=2 and col == 8:
                        temp_value["primary_fuel"] = str(value)
                    elif row != 1 and row !=2 and col == 9:
                        temp_value["other_fuel1"] = str(value)
                    elif row != 1 and row !=2 and col == 10:
                        temp_value["other_fuel2"] = str(value)
                    elif row != 1 and row !=2 and col == 11:
                        temp_value["other_fuel3"] = str(value)
                    elif row != 1 and row !=2 and col == 12:
                        temp_value["commissioning_year"] = str(value)
                    elif row != 1 and row !=2 and col == 13:
                        temp_value["owner"] = str(value)
                    elif row != 1 and row !=2 and col == 14:
                        temp_value["source"] = str(value)
                    elif row != 1 and row !=2 and col == 15:
                        temp_value["url"] = str(value)
                    elif row != 1 and row !=2 and col == 16:
                        temp_value["geolocation_source"] = str(value)
                    elif row != 1 and row !=2 and col == 17:
                        temp_value["wepp_id"] = str(value)
                    elif row != 1 and row !=2 and col == 18:
                        temp_value["year_of_capacity_data"] = str(value)
                    elif row != 1 and row !=2 and col == 19:
                        temp_value["generation_gwh_2013"] = str(value)
                    elif row != 1 and row !=2 and col == 20:
                        temp_value["generation_gwh_2014"] = str(value)
                    elif row != 1 and row !=2 and col == 21:
                        temp_value["generation_gwh_2015"] = str(value)
                    elif row != 1 and row !=2 and col == 22:
                        temp_value["generation_gwh_2016"] = str(value)
                    elif row != 1 and row !=2 and col == 23:
                        temp_value["generation_gwh_2017"] = str(value)
                    elif row != 1 and row !=2 and col == 24:
                        temp_value["estimated_generation_gwh"] = str(value)
                if "estimated_generation_gwh" in temp_value.keys():
                    temp_value["power_station_capacity_unit"] = "mw"
                    temp_values.append(temp_value)
            print(temp_values)
            for i in range(len(temp_values)):
                sql_str = "insert into fadianchang (country_iso3,country,power_station_name,power_station_code,\
                power_station_capacity,lat,lon,primary_fuel,other_fuel1,other_fuel2,other_fuel3,commissioning_year,\
                source,owner,url,geolocation_source,wepp_id,year_of_capacity_data,generation_gwh_2013,generation_gwh_2014,\
                generation_gwh_2015,generation_gwh_2016,generation_gwh_2017,estimated_generation_gwh,power_station_capacity_unit) values \
                ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" % \
                (temp_values[i]["country_iso3"], temp_values[i]["country"], temp_values[i]["power_station_name"],\
                temp_values[i]["power_station_code"],temp_values[i]["power_station_capacity"],temp_values[i]["lat"],\
                 temp_values[i]["lon"],temp_values[i]["primary_fuel"],temp_values[i]["other_fuel1"],temp_values[i]["other_fuel2"],\
                 temp_values[i]["other_fuel3"],temp_values[i]["commissioning_year"],temp_values[i]["source"],temp_values[i]["owner"],\
                 temp_values[i]["url"],temp_values[i]["geolocation_source"],temp_values[i]["wepp_id"],temp_values[i]["year_of_capacity_data"],\
                 temp_values[i]["generation_gwh_2013"],temp_values[i]["generation_gwh_2014"],temp_values[i]["generation_gwh_2015"],\
                 temp_values[i]["generation_gwh_2016"],temp_values[i]["generation_gwh_2017"],temp_values[i]["estimated_generation_gwh"],\
                 temp_values[i]["power_station_capacity_unit"])
                cursor.execute(sql_str)
                conn.commit()
            print(temp_file_path + "插入成功")
    conn.close()
def get_geojson(scale,zoom):
    lngDiff=0;
    latDiff=0;
    scaleCode='';
    if scale == 1000000:
        lngDiff=6
        latDiff=4
    elif scale == 500000:
        lngDiff=3
        latDiff=2
        scaleCode='B'
    elif scale == 250000:
        lngDiff=1.5
        latDiff=1
        scaleCode='C'
    elif scale==100000:
        lngDiff=0.5
        latDiff=1/3
        scaleCode='D'
    elif scale==50000:
        lngDiff=0.25
        latDiff=1/6
        scaleCode='E'
    elif scale==25000:
        lngDiff=0.125
        latDiff=1/12
        scaleCode='F'
    elif scale==10000:
        lngDiff=0.0625
        latDiff=1/24
        scaleCode='G'
    elif scale==5000:
        lngDiff=0.03125
        latDiff=1/48
        scaleCode='H'
    GridX0=-180
    GridX1=180
    GridY0=-88
    GridY1=88
    x0 = max(GridX0,73)
    y0 = max(GridY0,3)
    x1 = min(GridX1,136)
    y1 = min(GridY1,54)
    if (x1-x0)<lngDiff or (y1-y0)<latDiff:
        return None;
    features=[]
    #计算标准分幅网格行列范围
    col0=int((x0-GridX0)/lngDiff)
    col1=int((x1-GridX0)/lngDiff)
    row0=int((y0-GridY0)/latDiff)
    row1=int((y1-GridY0)/latDiff)
    millionRowCode='ABCDEFGHIJKLMNOPQRSTUV'
    for row0 in range(row1):
        gy0=GridY0+row0*latDiff
        gy1=gy0+latDiff
        gcy=(gy0+gy1)*0.5
        millionRow=int((gy0-0)/4)
        Hemisphere=''
        if millionRow<0:
            millionRow = -1-millionRow
            Hemisphere = 'S'
        for col0 in range(col1):
            gx0 = GridX0 + col0*lngDiff
            gx1 = gx0 + lngDiff
            gcx = (gx0+gx1)*0.5
            millionCol = int((gcx-GridX0)/6)+1
            coordinates=[[
                [gx0,gy0],
                [gx1,gy0],
                [gx1,gy1],
                [gx0,gy1],
                [gx0,gy0]
            ]];
            if millionCol<10:
                millionCol = '0'+str(millionCol)
            else:
                millionCol = millionCol
            gridID=str(Hemisphere)+str(millionRowCode[millionRow])+str(millionCol)
            if scaleCode!="":
                colID=int((fractional((gcx-GridX0)/6)*6)/lngDiff)+1
                rowID=int((fractional((GridY1-gcy)/4)*4)/latDiff)+1
                gridID+=scaleCode+formatInt(rowID,3)+formatInt(colID,3)
            feature={
                "type":"Feature",
                "geometry":{
                    "type":"Polygon",
                    "coordinates":coordinates
                },
                "properties":{
                    "id":gridID
                }
            }
            features.append(feature)
    data = {
        "type":"FeatureCollection",
        "features":features
    }
    with open("../public/showset/"+str(scale)+"-"+str(zoom)+".json","w",encoding="utf-8") as f:
        json.dump(data,f)
        print("已完成"+str(scale)+"比例尺分幅写入")
def fractional(x):
    x=abs(x)
    return x-math.floor(x)
def formatInt(x,length):
    result=''+str(x);
    length=length-len(result)
    while length>0:
        result = '0'+result
        length-=1
    return result
def test_get_geojson():
    temp=[
        10000,
        5000
    ]
    for i in range(len(temp)):
        get_geojson(temp[i],i+3)

def test_request():
    return1=os.system('ping 8.8.8.8')
    if return1:
        print('ping fail')
    else:
        print('ping ok')
import requests
def test_request2():
    proxies={"http":None,"https":None}
    url="http://webrd04.is.autonavi.com/appmaptile?lang=zh_cn&size=1&scale=1&style=8&x=46&y=23&z=6"
    result=requests.get(url,stream=True,proxies=proxies)
    print(result)


from PIL import Image
import numpy as np
# 高德,谷歌,OSM经纬度转墨卡托坐标系
def lnglat_to_Mercator(north_west):
    lng = north_west["lng"]*20037508.34/180
    temp_lat = math.log(math.tan((90 + north_west["lat"]) * math.pi / 360)) / (math.pi / 180)
    lat = temp_lat * 20037508.34 / 180
    return {
        "lng":lng,
        "lat":lat
    }
def test_get_file():
    file_paths=[
        {
            "url":"E:/SuperMapDownLoad/11/自绘多边形/15",
            "north_west":{
                "lat":39.32579941789298,
                "lng":114.99689162131573
            },
            "resolution": -0.26687106258428456,
            "zoom":"15"
        }
    ]
    file_paths2 = [
        {
            "url": "E:/SuperMapDownLoad/2/自绘多边形/8",
            "north_west": {
                "lat": 31.06401549873301,
                "lng": 100.57002337462964
            },
            "resolution": 574.0211126682296,
            "zoom": "8"
        }
    ]
    # 遍历文件地址集合
    for s in range(len(file_paths)):
        new_image_path=file_paths[s]["url"]+"/"+file_paths[s]["zoom"]+".jpg"
        new_tif_path=file_paths[s]["url"]+"/"+file_paths[s]["zoom"]+".tif"
        new_image_prj=file_paths[s]["url"]+"/"+file_paths[s]["zoom"]+".prj"
        new_image_tfw=file_paths[s]["url"]+"/"+file_paths[s]["zoom"]+".jgw"
        resolution = file_paths[s]["resolution"]
        north_west  = lnglat_to_Mercator(file_paths[s]["north_west"])
        print(north_west)
        images=[]
        # 获取瓦片拼接存储地址
        for item in os.listdir(file_paths[s]["url"]):
            temp_images=[]
            if os.path.isdir(file_paths[s]["url"]+"/"+item):
                for item2 in os.listdir(file_paths[s]["url"]+"/"+item):
                    print(file_paths[s]["url"]+"/"+item+"/"+item2)
                    image_content=Image.open(file_paths[s]["url"]+"/"+item+"/"+item2)
                    temp_images.append(image_content)
                # 瓦片拼接 如果是百度地图则从下至上 否则从上至下
                # temp_images.reverse()
                images.append(temp_images)
        pass
        # 获取图片总宽度和高度
        total_width=len(images)*256
        total_height=len(images[0])*256
        new_image=Image.new("RGB",(total_width,total_height))
        for f in range(len(images)):
            # x偏移量
            x_off=f*256
            for c in range(len(images[f])):
                #y偏移量
                y_off=c*256
                new_image.paste(images[f][c],(x_off,y_off))
        # 参数quality将影响图片质量95为最好质量
        new_image.save(new_image_path,quality=95)
        # 创建tfw文件 用于标识tif的位置
        fd=open(new_image_tfw,mode="w",encoding="utf-8")
        # 写入x方向 像素分辨率
        fd.write(str(resolution)+'\r')
        # 写入平移量
        fd.write('0.0000000000\r')
        # 写入旋转角度
        fd.write('0.0000000000\r')
        # 写入y方向 像素分辨率
        fd.write(str(resolution)+'\r')
        # 写入图像左上角x坐标
        fd.write(str(north_west["lng"])+'\r')
        # 写入图像左上角y坐标
        fd.write(str(north_west["lat"])+'\r')
        fd.close()

        # 生成坐标系文件
        # 创建prj文件 用于标识tif的位置
        prj=open(new_image_prj,mode="w",encoding="utf-8")
        # 坐标系字符串
        prj.write(
            'PROJCS["WGS_1984_Web_Mercator_Auxiliary_Sphere",GEOGCS["GCS_WGS_1984",DATUM["D_WGS_1984",SPHEROID["WGS_1984",6378137.0,298.257223563]],PRIMEM["Greenwich",0.0],UNIT["Degree",0.017453292519943295]],PROJECTION["Mercator_Auxiliary_Sphere"],PARAMETER["False_Easting",0.0],PARAMETER["False_Northing",0.0],PARAMETER["Central_Meridian",0.0],PARAMETER["Standard_Parallel_1",0.0],PARAMETER["Auxiliary_Sphere_Type",0.0],UNIT["Meter",1.0],EXTENSION["PROJ4","+proj=merc +a=6378137 +b=6378137 +lat_ts=0.0 +lon_0=0.0 +x_0=0.0 +y_0=0 +k=1.0 +units=m +nadgrids=@null +wktext  +no_defs"]]')
        prj.close()
        driver=gdal.GetDriverByName('GTiff')
        in_ds=gdal.Open(new_image_path)
        #  获取仿射矩阵信息,投影信息
        im_geotrans=in_ds.GetGeoTransform()
        print(im_geotrans)
        im_proj=in_ds.GetProjection()
        data_width=in_ds.RasterXSize
        data_height=in_ds.RasterYSize
        im_data=in_ds.ReadAsArray(0,0,data_width,data_height)
        if 'int8' in im_data.dtype.name:
            datatype=gdal.GDT_Byte
        elif 'int16' in im_data.dtype.name:
            datatype=gdal.GDT_UInt16
        else:
            datatype=gdal.GDT_Float32
        if len(im_data.shape)==2:
            im_data=np.array([im_data])
        im_bands, im_height, im_width = im_data.shape
        # 创建tif文件
        out_ds=driver.Create(new_tif_path,im_width,im_height,im_bands,datatype)
        out_ds.SetGeoTransform(im_geotrans)  # 写入仿射变换参数
        out_ds.SetProjection(im_proj)  # 写入投影
        for c in range(im_bands):
            out_ds.GetRasterBand(c+1).WriteArray(im_data[c])
        # in_files =[]
            # print(root)
        # in_fn=in_files[0]
        # #获取待镶嵌栅格的最大最小的坐标值
        # min_x,max_y,max_x,min_y=GetExtent(in_fn)
        # for in_fn in in_files[1:]:
        #     minx,maxy,maxx,miny=GetExtent(in_fn)
        #     min_x=min(min_x,minx)
        #     min_y=min(min_y,miny)
        #     max_x=max(max_x,maxx)
        #     max_y=max(max_y,maxy)
        # #计算镶嵌后影像的行列号
        # in_ds=gdal.Open(in_files[0])
        # in_ds.SetProjection("EPSG:3857")
        # geotrans=list(in_ds.GetGeoTransform())
        # width=geotrans[1]
        # height=geotrans[5]
        # columns=math.ceil((max_x-min_x)/width)
        # rows=math.ceil((max_y-min_y)/(-height))
        # print(columns)
        # print(rows)
        # in_band=in_ds.GetRasterBand(1)
        # print(in_band)
        # driver=gdal.GetDriverByName('GTiff')
        # if os.path.exists(file_paths[s]+'/mosaiced_image.tif'):
        #     os.remove(file_paths[s]+'/mosaiced_image.tif')
        # out_ds=driver.Create(file_paths[s]+'/mosaiced_image.tif',768,768,1,in_band.DataType)
        # out_ds.SetProjection("EPSG:3857")
        # geotrans[0]=min_x
        # geotrans[3]=max_y
        # out_ds.SetGeoTransform(geotrans)
        # out_band=out_ds.GetRasterBand(1)
        # #定义仿射逆变换
        # inv_geotrans=gdal.InvGeoTransform(geotrans)
        # #开始逐渐写入
        # for in_fns in in_files:
        #     in_ds2=gdal.Open(in_fns)
        #     in_ds2.SetProjection("EPSG:4326")
        #     in_gt=in_ds2.GetGeoTransform()
        #     #仿射逆变换
        #     offset=gdal.ApplyGeoTransform(inv_geotrans,in_gt[0],in_gt[3])
        #     x,y=map(int,offset)
        #     # print(x,y)
        #     # trans=gdal.Transformer(in_ds2,out_ds,[])  #in_ds是源栅格，out_ds是目标栅格
        #     # success,xyz=trans.TransformPoint(False,0,0)  #计算in_ds中左上角像元对应out_ds中的行列号
        #     # x,y,z=map(int,xyz)
        #     # print(x,y,z)
        #     data=in_ds2.GetRasterBand(1).ReadAsArray()
        #     out_band.WriteArray(data,0,1)  #x，y是开始写入时左上角像元行列号
        # del in_ds,out_band,out_ds

#获取影像的左上角和右下角坐标
def GetExtent(in_fn):
    ds=gdal.Open(in_fn)
    ds.SetProjection("EPSG:4326")
    geotrans=list(ds.GetGeoTransform())
    xsize=ds.RasterXSize
    ysize=ds.RasterYSize
    min_x=geotrans[0]
    max_y=geotrans[3]
    max_x=geotrans[0]+xsize*geotrans[1]
    min_y=geotrans[3]+ysize*geotrans[5]
    ds=None
    return min_x,max_y,max_x,min_y
test_get_file()